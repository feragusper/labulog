import csv
import io
import re
import unicodedata
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlmodel import Session, select

from ..crud import get_or_create_company, upsert_posting
from ..db import get_session
from ..deps import get_current_user
from ..models import AppStatus, Application, JobPosting, Priority, StatusEvent, User, utcnow
from ..schemas import PostingCreate

router = APIRouter(prefix="/api/import", tags=["import"])

# CSV stage column -> the status that reaching it implies.
STAGE_COLUMNS = [
    ("Apply", AppStatus.applied),
    ("First Contact", AppStatus.first_contact),
    ("Screening", AppStatus.screening),
    ("Technical I", AppStatus.technical_interview),
    ("Technical II", AppStatus.technical_interview),
    ("Manager I", AppStatus.manager_interview),
    ("Manager II", AppStatus.manager_interview),
    ("Proposal", AppStatus.proposal),
]

SKIP_TOKENS = {"", "-", "waiting", "?", "`", "n/a"}


class PendingPosting(BaseModel):
    url: Optional[str] = None
    title: str
    company_name: str
    location: Optional[str] = None
    country: Optional[str] = None
    industry: Optional[str] = None
    salary_min: Optional[int] = None
    salary_max: Optional[int] = None
    currency: Optional[str] = None
    source: Optional[str] = None


class PendingRow(BaseModel):
    """A row the importer skipped, handed back so the user can review it and
    decide to force-add it or discard it from the app."""
    reason: str
    posting: PendingPosting
    status: AppStatus = AppStatus.applied
    priority: Optional[Priority] = None
    applied_at: Optional[datetime] = None
    follow_up_date: Optional[datetime] = None
    notes: Optional[str] = None


class ImportResult(BaseModel):
    imported: int
    skipped: int
    errors: List[str]
    pending: List[PendingRow] = []


# Excel stores dates as a serial day count from this epoch (the 1900 system).
_EXCEL_EPOCH = datetime(1899, 12, 30)

# Day-first (dd/mm/yyyy) is prioritised over US mm/dd, so ambiguous dates like
# 03/04/2024 read as 3 April, not 4 March.
_DATE_FORMATS = (
    "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y/%m/%d",
    "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d.%m.%Y",
    "%m/%d/%Y", "%m/%d/%y",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y", "%d-%b-%y",
)


def _parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    # openpyxl hands back real date/datetime objects; use them as-is.
    if isinstance(value, datetime):
        return value
    from datetime import date
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    v = str(value).strip()
    if not v or v.lower() in SKIP_TOKENS:
        return None

    # A bare number in a plausible range is an Excel serial date, not a year.
    if re.fullmatch(r"\d{4,6}(\.0+)?", v):
        n = int(float(v))
        if 20000 <= n <= 80000:  # ~1954 .. ~2119
            return _EXCEL_EPOCH + timedelta(days=n)

    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue

    # Last resort: dateutil copes with month names and odd separators. Translate
    # Spanish month names first (dateutil only knows English).
    try:
        from dateutil import parser as _dtp
        return _dtp.parse(_es_months_to_en(v), dayfirst=True, fuzzy=True)
    except Exception:
        return None


_ES_MONTHS = {
    "enero": "January", "febrero": "February", "marzo": "March", "abril": "April",
    "mayo": "May", "junio": "June", "julio": "July", "agosto": "August",
    "septiembre": "September", "setiembre": "September", "octubre": "October",
    "noviembre": "November", "diciembre": "December",
    "ene": "Jan", "feb": "Feb", "mar": "Mar", "abr": "Apr", "jun": "Jun",
    "jul": "Jul", "ago": "Aug", "sep": "Sep", "set": "Sep", "oct": "Oct",
    "nov": "Nov", "dic": "Dec",
}


def _es_months_to_en(text: str) -> str:
    # Strip accents so "años"/"días" noise and accented months normalise, then
    # swap Spanish month words for English ones.
    decomposed = unicodedata.normalize("NFKD", text.lower())
    flat = "".join(c for c in decomposed if not unicodedata.combining(c))
    return re.sub(r"[a-z]+", lambda m: _ES_MONTHS.get(m.group(0), m.group(0)), flat)


def _parse_money(value: Optional[str]):
    """'€50,000.00' -> (50000, 'EUR'). Returns (amount|None, currency|None)."""
    if not value or value.strip() in SKIP_TOKENS:
        return None, None
    currency = "EUR" if "€" in value else "USD" if "$" in value else None
    digits = re.sub(r"[^\d]", "", value.split(".")[0])  # drop decimals + symbols
    return (int(digits) if digits else None), currency


def _infer_final_status(result_text: str, reached: AppStatus) -> AppStatus:
    """Result free-text overrides the furthest stage reached when it signals an outcome."""
    t = (result_text or "").lower()
    if "pagaron" in t or "pagó" in t:
        return AppStatus.offer
    if "rechaz" in t or "no go" in t or "no van a seguir" in t or "no aceptaron" in t or "descartó" in t:
        return AppStatus.rejected
    if "on hold" in t or "paralizada" in t or "cambié" in t or "pasaron a" in t:
        return AppStatus.withdrawn
    if "baja" in t or "nunca hubo respuesta" in t or "ni pelota" in t or "ni responden" in t or "ni bola" in t:
        # If we got past applied, it's a real ghost; otherwise also ghosted.
        return AppStatus.ghosted
    return reached


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-") or "row"


@router.post("/csv", response_model=ImportResult)
async def import_csv(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current: User = Depends(get_current_user),
):
    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    required = {"Company", "Apply"}
    if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
        raise HTTPException(
            status_code=400,
            detail="CSV needs at least 'Company' and 'Apply' columns (job-scouting format).",
        )

    imported = 0
    skipped = 0
    errors: List[str] = []

    for idx, row in enumerate(reader, start=2):  # row 1 is the header
        company = (row.get("Company") or "").strip()
        if not company:
            skipped += 1
            continue

        client = (row.get("Company 2") or "").strip()
        result_text = (row.get("Result") or "").strip()
        title = f"{company} · {client}" if client and client != company else company

        # Synthetic, unique, stable URL (re-import upserts the same posting).
        url = f"imported://job-scouting/{idx}-{_slug(company)}"

        try:
            applied_dt = _parse_date(row.get("Apply"))
            salary, currency = _parse_money(row.get("Money"))

            # Build the timeline from whatever stage dates exist.
            events: List[tuple] = []
            reached = AppStatus.applied
            for col, status in STAGE_COLUMNS:
                dt = _parse_date(row.get(col))
                if dt:
                    events.append((dt, status))
                    reached = status
            if not events:
                events.append((applied_dt or utcnow(), AppStatus.applied))

            final_status = _infer_final_status(result_text, reached)

            posting = session.exec(select(JobPosting).where(JobPosting.url == url)).first()
            if not posting:
                comp = get_or_create_company(session, company)
                posting = JobPosting(
                    company_id=comp.id, title=title, url=url, source="import",
                    salary_min=salary, salary_max=salary, currency=currency,
                )
                session.add(posting)
                session.commit()
                session.refresh(posting)

            exists = session.exec(
                select(Application).where(
                    Application.posting_id == posting.id,
                    Application.user_id == current.id,
                )
            ).first()
            if exists:
                skipped += 1
                continue

            app = Application(
                user_id=current.id, posting_id=posting.id, status=final_status,
                applied_at=applied_dt or utcnow(), channel=None,
                notes=result_text or None,
            )
            session.add(app)
            session.commit()
            session.refresh(app)

            # Timeline events (sorted), plus a terminal event if the outcome
            # differs from the furthest stage reached.
            events.sort(key=lambda e: e[0])
            for dt, status in events:
                session.add(StatusEvent(application_id=app.id, status=status, at=dt))
            if final_status != reached:
                session.add(StatusEvent(application_id=app.id, status=final_status,
                                        at=events[-1][0], note="inferido del resultado"))
            session.commit()
            imported += 1
        except Exception as e:  # keep going; report the bad row
            session.rollback()
            errors.append(f"fila {idx} ({company}): {e}")

    return ImportResult(imported=imported, skipped=skipped, errors=errors)


# ---------------------------------------------------------------------------
# General importer: mirrors the export.csv format (and any spreadsheet close to
# it). Maps whatever columns it recognises, leaves the rest blank, and lets the
# user finish editing each row in the app afterwards.
# ---------------------------------------------------------------------------

# canonical field -> list of accepted header names (normalised: lowercased,
# non-alphanumerics stripped, so "Applied At", "applied_at" and "fecha" all fit).
FIELD_ALIASES: Dict[str, List[str]] = {
    "company": ["company", "empresa", "companyname", "compania"],
    "title": ["title", "role", "position", "puesto", "cargo", "titulo", "rol"],
    "status": ["status", "estado"],
    "priority": ["priority", "prioridad"],
    "applied_at": ["appliedat", "applied", "applieddate", "apply", "fecha",
                   "fechapostulacion", "date", "appliedon",
                   "aplicada", "aplicado", "aplicacion", "aplico",
                   "fechaaplicacion", "fechaaplicada", "fechadeaplicacion",
                   "postulada", "postulado", "fechapostulada",
                   "fechadepostulacion"],
    "follow_up_date": ["followupdate", "followup", "seguimiento",
                       "fechaseguimiento", "proximopaso"],
    "salary_min": ["salarymin", "salariomin", "minsalary", "sueldomin"],
    "salary_max": ["salarymax", "salariomax", "maxsalary", "sueldomax"],
    "salary": ["salary", "salario", "sueldo", "money", "compensacion"],
    "currency": ["currency", "moneda"],
    "source": ["source", "fuente", "origen", "platform", "plataforma", "channel"],
    "url": ["url", "link", "enlace"],
    "location": ["location", "ubicacion", "ciudad", "city"],
    "country": ["country", "pais"],
    "industry": ["industry", "industria", "sector", "rubro"],
    "notes": ["notes", "notas", "note", "nota", "comentarios", "comments",
              "result", "resultado"],
}

# Loose value -> AppStatus. Checked as substrings against the normalised cell.
STATUS_SYNONYMS: List[Tuple[str, AppStatus]] = [
    ("firstcontact", AppStatus.first_contact),
    ("primercontacto", AppStatus.first_contact),
    ("screening", AppStatus.screening),
    ("technical", AppStatus.technical_interview),
    ("tecnica", AppStatus.technical_interview),
    ("tecnico", AppStatus.technical_interview),
    ("manager", AppStatus.manager_interview),
    ("interview", AppStatus.interview),
    ("entrevista", AppStatus.interview),
    ("proposal", AppStatus.proposal),
    ("propuesta", AppStatus.proposal),
    ("offer", AppStatus.offer),
    ("oferta", AppStatus.offer),
    ("accepted", AppStatus.accepted),
    ("aceptad", AppStatus.accepted),
    ("rejected", AppStatus.rejected),
    ("rechazad", AppStatus.rejected),
    ("cancelled", AppStatus.cancelled),
    ("cancelad", AppStatus.cancelled),
    ("ghosted", AppStatus.ghosted),
    ("ghost", AppStatus.ghosted),
    ("withdrawn", AppStatus.withdrawn),
    ("retirad", AppStatus.withdrawn),
    ("saved", AppStatus.saved),
    ("guardad", AppStatus.saved),
    ("applied", AppStatus.applied),
    ("postulad", AppStatus.applied),
    ("aplicad", AppStatus.applied),
]

PRIORITY_SYNONYMS: List[Tuple[str, Priority]] = [
    ("high", Priority.high), ("alta", Priority.high),
    ("medium", Priority.medium), ("media", Priority.medium),
    ("low", Priority.low), ("baja", Priority.low),
]


def _norm_key(text: str) -> str:
    # Lowercase, strip accents (á->a, ó->o, ñ->n) and drop non-alphanumerics,
    # so "Fecha de aplicación", "aplicada" and "applied_at" all collapse to a
    # comparable key.
    lowered = (text or "").lower()
    decomposed = unicodedata.normalize("NFKD", lowered)
    no_accents = "".join(c for c in decomposed if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", "", no_accents)


def _build_header_map(fieldnames: List[str]) -> Dict[str, str]:
    """Map canonical field -> the original header that supplies it."""
    normalised = {_norm_key(h): h for h in fieldnames if h}
    mapping: Dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                mapping[field] = normalised[alias]
                break
    return mapping


def _cell(row: dict, header_map: Dict[str, str], field: str) -> Optional[str]:
    header = header_map.get(field)
    if not header:
        return None
    value = row.get(header)
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in SKIP_TOKENS:
        return None
    return text


def _parse_status(value: Optional[str]) -> AppStatus:
    if not value:
        return AppStatus.applied
    key = _norm_key(value)
    try:
        return AppStatus(value.strip().lower())
    except ValueError:
        pass
    for token, status in STATUS_SYNONYMS:
        if token in key:
            return status
    return AppStatus.applied


def _parse_priority(value: Optional[str]) -> Optional[Priority]:
    if not value:
        return None
    key = _norm_key(value)
    for token, prio in PRIORITY_SYNONYMS:
        if token in key:
            return prio
    return None


def _parse_int(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    digits = re.sub(r"[^\d]", "", value.split(".")[0].split(",")[0])
    return int(digits) if digits else None


def _read_table(filename: str, raw: bytes) -> Tuple[List[str], List[dict]]:
    """Return (fieldnames, rows) from a CSV or Excel upload."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(raw)
    if name.endswith(".xls"):
        return _read_xls(raw)
    # default: CSV / text
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), list(reader)


def _read_xlsx(raw: bytes) -> Tuple[List[str], List[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(status_code=400,
                            detail="Soporte de .xlsx no disponible en el servidor.")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows = []
    for values in rows_iter:
        rows.append({headers[i]: values[i] for i in range(len(headers)) if i < len(values)})
    return headers, rows


def _read_xls(raw: bytes) -> Tuple[List[str], List[dict]]:
    try:
        import xlrd
    except ImportError:  # pragma: no cover
        raise HTTPException(
            status_code=400,
            detail="Soporte de .xls no disponible; guardá el archivo como .xlsx o .csv.",
        )
    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], []
    headers = [str(c.value).strip() for c in sheet.row(0)]
    rows = []
    for r in range(1, sheet.nrows):
        cells = sheet.row(r)
        rows.append({headers[i]: cells[i].value for i in range(len(headers)) if i < len(cells)})
    return headers, rows


@router.post("/applications", response_model=ImportResult)
async def import_applications(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current: User = Depends(get_current_user),
):
    raw = await file.read()
    try:
        fieldnames, table = _read_table(file.filename or "", raw)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    header_map = _build_header_map(fieldnames)
    if "company" not in header_map and "title" not in header_map:
        raise HTTPException(
            status_code=400,
            detail="El archivo necesita al menos una columna de empresa o de puesto.",
        )

    imported = 0
    skipped = 0
    errors: List[str] = []
    pending: List[PendingRow] = []

    for idx, row in enumerate(table, start=2):  # row 1 is the header
        company = _cell(row, header_map, "company")
        title = _cell(row, header_map, "title")
        if not company and not title:
            skipped += 1  # blank row, nothing to review
            continue
        company = company or "(sin empresa)"
        title = title or "(sin título)"

        try:
            url = _cell(row, header_map, "url")
            # Give url-less rows a stable synthetic key so re-imports upsert
            # instead of duplicating.
            if not url:
                url = f"imported://app/{_slug(company)}-{_slug(title)}"

            salary_min = _parse_int(_cell(row, header_map, "salary_min"))
            salary_max = _parse_int(_cell(row, header_map, "salary_max"))
            single = _parse_int(_cell(row, header_map, "salary"))
            if single and not salary_min and not salary_max:
                salary_min = salary_max = single

            status = _parse_status(_cell(row, header_map, "status"))
            applied_at = _parse_date(_cell(row, header_map, "applied_at"))
            follow_up = _parse_date(_cell(row, header_map, "follow_up_date"))
            priority = _parse_priority(_cell(row, header_map, "priority"))
            notes = _cell(row, header_map, "notes")
            location = _cell(row, header_map, "location")
            country = _cell(row, header_map, "country")
            industry = _cell(row, header_map, "industry")
            currency = _cell(row, header_map, "currency")
            source = _cell(row, header_map, "source") or "import"

            posting = upsert_posting(session, PostingCreate(
                url=url, title=title, company_name=company,
                location=location, country=country, industry=industry,
                salary_min=salary_min, salary_max=salary_max,
                currency=currency, source=source,
            ))

            exists = session.exec(
                select(Application).where(
                    Application.posting_id == posting.id,
                    Application.user_id == current.id,
                )
            ).first()
            if exists:
                # Don't touch the existing application (may hold manual edits).
                # Hand the row back so the user can review the conflict.
                skipped += 1
                pending.append(PendingRow(
                    reason="already_exists",
                    posting=PendingPosting(
                        url=url, title=title, company_name=company,
                        location=location, country=country, industry=industry,
                        salary_min=salary_min, salary_max=salary_max,
                        currency=currency, source=source,
                    ),
                    status=status, priority=priority,
                    applied_at=applied_at, follow_up_date=follow_up, notes=notes,
                ))
                continue

            app = Application(
                user_id=current.id, posting_id=posting.id, status=status,
                priority=priority, follow_up_date=follow_up,
                applied_at=applied_at or utcnow(), notes=notes,
            )
            session.add(app)
            session.commit()
            session.refresh(app)

            session.add(StatusEvent(application_id=app.id, status=status,
                                    at=applied_at or utcnow()))
            session.commit()
            imported += 1
        except Exception as e:  # keep going; report the bad row
            session.rollback()
            errors.append(f"fila {idx} ({company}): {e}")

    return ImportResult(imported=imported, skipped=skipped, errors=errors, pending=pending)
